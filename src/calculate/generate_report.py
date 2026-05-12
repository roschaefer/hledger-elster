from __future__ import annotations

import csv
import io
import re
import sys
import unicodedata
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from calculate.report.est import est_rows
from calculate.report.euer import euer_rows
from calculate.report.herleitung import FORM_KEYS, TrailSheet, herleitung_sheets
from calculate.report.ust import ust_rows
from config import load_config
from ingest.enrich import build_dataset
from paths import elster_config_path, ledger_journal_path, tax_data_dir

# Fixed timestamps so repeated runs produce byte-identical files.
_ZIP_DATE_TIME = (2000, 1, 1, 0, 0, 0)
_CORE_XML = (
    b"<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\r\n"
    b"<cp:coreProperties"
    b' xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"'
    b' xmlns:dc="http://purl.org/dc/elements/1.1/"'
    b' xmlns:dcterms="http://purl.org/dc/terms/"'
    b' xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
    b"<dc:creator>openpyxl</dc:creator>"
    b'<dcterms:created xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:created>'
    b'<dcterms:modified xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>'
    b"</cp:coreProperties>"
)

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
_DEFAULT_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
_DEFAULT_FONT = Font(color="000000")
_HEADER_FONT = Font(bold=True, color="000000")
_BOLD_FONT = Font(bold=True, color="000000")
_BLANK_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
_SECTION_FILL = PatternFill(fill_type="solid", fgColor="E9EFF7")
_SUBTOTAL_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
_TOTAL_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")

# Maps form key → (XLSX tab name, CSV filename stem)
_ELSTER_FORMS = {
    "einnahmen-ueberschuss-rechnung": ("EÜR", "einnahmen-ueberschuss-rechnung"),
    "umsatzsteuer": ("USt", "umsatzsteuer"),
    "einkommensteuer": ("ESt", "einkommensteuer"),
}
_INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")
_NON_ALNUM_FILENAME_CHARS = re.compile(r"[^a-z0-9]+")


# ── ZIP stabilisation ─────────────────────────────────────────────────────────


def _stabilize_zip(path: Path) -> None:
    data = path.read_bytes()
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data), "r") as zin:
        entries = sorted(zin.infolist(), key=lambda e: e.filename)
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for entry in entries:
                info = zipfile.ZipInfo(entry.filename, date_time=_ZIP_DATE_TIME)
                info.compress_type = zipfile.ZIP_DEFLATED
                content = _CORE_XML if entry.filename == "docProps/core.xml" else zin.read(entry.filename)
                zout.writestr(info, content)
    path.write_bytes(buf.getvalue())


# ── CSV writers ───────────────────────────────────────────────────────────────


def _write_rows_csv(path: Path, rows: list[dict[str, str]], touched_files: set[Path]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        return
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()), lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    touched_files.add(path.resolve())


def _write_trail_csv(path: Path, sheet: TrailSheet, touched_files: set[Path]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh, lineterminator="\n")
        writer.writerow(sheet.headers)
        for row in sheet.rows:
            writer.writerow(row.cells)
    touched_files.add(path.resolve())


def _tab_csv_name(sheet_name: str) -> str:
    normalized = unicodedata.normalize("NFKD", sheet_name)
    ascii_name = normalized.encode("ascii", "ignore").decode("ascii").lower()
    stem = _NON_ALNUM_FILENAME_CHARS.sub("-", ascii_name).strip("-")
    return f"{stem or 'sheet'}.csv"


def _xlsx_sheet_title(sheet_name: str) -> str:
    title = _INVALID_SHEET_TITLE_CHARS.sub("-", sheet_name).strip()
    return (title or "Sheet")[:31]


# ── XLSX writers ──────────────────────────────────────────────────────────────


def _is_blank(row: dict[str, str]) -> bool:
    return all(v == "" for v in row.values())


def _is_section_header(row: dict[str, str]) -> bool:
    return row.get("Kennzahl", "").startswith("# ")


def _apply_default_light_style(cell) -> None:
    cell.font = _DEFAULT_FONT
    cell.fill = _DEFAULT_FILL


def _style_row_light(row) -> None:
    for cell in row:
        _apply_default_light_style(cell)


def _write_summary_sheet(ws: Worksheet, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        if _is_section_header(row):
            display = dict(row)
            display["Kennzahl"] = row["Kennzahl"][2:]  # strip "# "
            ws.append([display.get(h, "") for h in headers])
            _style_row_light(ws[ws.max_row])
            for cell in ws[ws.max_row]:
                cell.font = _BOLD_FONT
                cell.fill = _SECTION_FILL
        elif _is_blank(row):
            ws.append([row.get(h, "") for h in headers])
            _style_row_light(ws[ws.max_row])
            for cell in ws[ws.max_row]:
                cell.fill = _BLANK_FILL
        else:
            ws.append([row.get(h, "") for h in headers])
            _style_row_light(ws[ws.max_row])

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(row.get(header, "")) for row in rows), default=0)
        max_len = max(max_len, len(header))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def _write_ust_sheet(ws: Worksheet, rows: list[dict[str, str]]) -> None:
    """Writer for the vertical USt layout.

    Row type is inferred from the Zeitraum column:
      - "YYYY-MM"   → monthly   (no fill)
      - "YYYY QN"   → quarterly → subtotal fill + bold
      - "YYYY"      → annual    → total fill + bold
      - all empty   → blank separator
    """
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])
        row_num = ws.max_row
        _style_row_light(ws[row_num])
        zeitraum = row.get("Zeitraum", "")
        if not any(row.values()):
            for cell in ws[row_num]:
                cell.fill = _BLANK_FILL
        elif "Q" in zeitraum:
            for cell in ws[row_num]:
                cell.fill = _SUBTOTAL_FILL
                cell.font = _BOLD_FONT
        elif "-" not in zeitraum and zeitraum:
            for cell in ws[row_num]:
                cell.fill = _TOTAL_FILL
                cell.font = _BOLD_FONT

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(row.get(header, "")) for row in rows), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_len, len(header)) + 2, 50)


def _write_trail_sheet(ws: Worksheet, sheet: TrailSheet) -> None:
    ws.append(sheet.headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for trail_row in sheet.rows:
        ws.append(trail_row.cells)
        row_num = ws.max_row
        _style_row_light(ws[row_num])
        if trail_row.bold:
            for cell in ws[row_num]:
                cell.font = _BOLD_FONT
        if trail_row.fill == "subtotal":
            for cell in ws[row_num]:
                cell.fill = _SUBTOTAL_FILL
        elif trail_row.fill == "total":
            for cell in ws[row_num]:
                cell.fill = _TOTAL_FILL

    for col_idx, header in enumerate(sheet.headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(row.cells[col_idx - 1]) for row in sheet.rows if col_idx - 1 < len(row.cells)),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len, len(header)) + 2, 50)


def _save_xlsx(path: Path, wb: openpyxl.Workbook, touched_files: set[Path]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    _stabilize_zip(path)
    touched_files.add(path.resolve())


def _unlink_if_exists(path: Path, touched_files: set[Path]) -> None:
    if path.exists():
        path.unlink()
        touched_files.add(path.resolve())


def _warn_about_untouched_files(data_dir: Path, touched_files: set[Path]) -> None:
    if not data_dir.exists():
        return
    untouched = sorted(
        path.resolve() for path in data_dir.rglob("*") if path.is_file() and path.resolve() not in touched_files
    )
    if not untouched:
        return

    print(file=sys.stderr)
    print(
        "Warning: untouched files remain in ELSTER export directory. "
        "Consider emptying the export directory and running the tool again.",
        file=sys.stderr,
    )
    for path in untouched:
        print(f"  {path}", file=sys.stderr)
    print(file=sys.stderr)


# ── main ──────────────────────────────────────────────────────────────────────


def main() -> int:
    journal_path = ledger_journal_path()
    data_dir = tax_data_dir()
    config = load_config(elster_config_path())
    touched_files: set[Path] = set()

    if not journal_path.exists():
        print(f"Journal not found: {journal_path}", file=sys.stderr)
        return 1

    dataset = build_dataset(journal_path)
    years = sorted({p.year for p in dataset})

    for year in years:
        base = data_dir / str(year)
        print(f"  {year}")

        # ── steuererklaerung.xlsx + steuererklaerung/ CSVs ───────────────
        elster_rows = {
            "EÜR": euer_rows(dataset, year, config),
            "USt": ust_rows(dataset, year),
            "ESt": est_rows(dataset, year),
        }
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for tab_name, rows in elster_rows.items():
            ws = wb.create_sheet(title=tab_name)
            if tab_name == "USt":
                _write_ust_sheet(ws, rows)
            else:
                _write_summary_sheet(ws, rows)
        _save_xlsx(base / "steuererklaerung.xlsx", wb, touched_files)

        for form_key, (tab_name, csv_stem) in _ELSTER_FORMS.items():
            _write_rows_csv(base / "steuererklaerung" / f"{csv_stem}.csv", elster_rows[tab_name], touched_files)

        # ── herleitung/ ───────────────────────────────────────────────────
        all_herleitung = herleitung_sheets(dataset, year)

        for form_key in FORM_KEYS:
            sheets = all_herleitung.get(form_key, [])
            if not sheets:
                continue
            _, csv_stem = _ELSTER_FORMS[form_key]

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for sheet in sheets:
                ws = wb.create_sheet(title=_xlsx_sheet_title(sheet.name))
                _write_trail_sheet(ws, sheet)
            _save_xlsx(base / "herleitung" / f"{csv_stem}.xlsx", wb, touched_files)

            for sheet in sheets:
                _write_trail_csv(base / "herleitung" / csv_stem / _tab_csv_name(sheet.name), sheet, touched_files)

        ignored_sheets = all_herleitung.get("ignoriert", [])
        if ignored_sheets:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for sheet in ignored_sheets:
                ws = wb.create_sheet(title=_xlsx_sheet_title(sheet.name))
                _write_trail_sheet(ws, sheet)
            _save_xlsx(base / "herleitung" / "ignoriert.xlsx", wb, touched_files)
            for sheet in ignored_sheets:
                _write_trail_csv(base / "herleitung" / _tab_csv_name(sheet.name), sheet, touched_files)
        else:
            _unlink_if_exists(base / "herleitung" / "ignoriert.xlsx", touched_files)
            _unlink_if_exists(base / "herleitung" / "ignoriert.csv", touched_files)

    _warn_about_untouched_files(data_dir, touched_files)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (ValueError, FileNotFoundError) as exc:
        raise SystemExit(f"Error: {exc}") from None
